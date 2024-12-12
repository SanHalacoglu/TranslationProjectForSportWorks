import os
import threading
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import utils
from openai import OpenAI
from utils import save_excel_template


class TranslationApp:
    def __init__(self, root):
        self.root = root
        self.root.title("English to Turkish Translator For Product Description")

        # UI References
        self.template_label = None
        self.template_button = None
        self.input_label = None
        self.file_label = None
        self.upload_button = None
        self.input_file_path = None
        self.api_key = None
        self.entry_api_key = None
        self.process_log_label = None
        self.process_button = None
        self.abort_button = None

        # Threading and Abort
        self.processing_thread = None
        self.abort_requested_state = False

        # Initialize UI
        self.create_template_section()
        self.create_input_section()
        self.create_processing_section()
        self.create_abort_section()

    def create_template_section(self):
        """Create UI elements to download the template file."""
        self.template_label = tk.Label(
            self.root,
            text="Please look at the template file for formatting:",
            bg="white",
            fg="black",
            anchor="w"
        )
        self.template_label.grid(row=0, column=0, padx=5, pady=5, sticky="ew")

        self.template_button = tk.Button(
            self.root,
            text="Download Template",
            command=self.download_template_file
        )
        self.template_button.grid(row=0, column=2, padx=5, pady=5, sticky="ew")

    def create_input_section(self):
        """Create UI elements for selecting the input Excel file."""
        self.input_label = tk.Label(
            self.root,
            text="Please upload the Excel file you want to translate:",
            bg="white",
            fg="black",
            anchor="w"
        )
        self.input_label.grid(row=1, column=0, padx=5, pady=5, sticky="ew")

        self.file_label = tk.Label(
            self.root,
            text="No file selected",
            bg="white",
            fg="black",
            width=30,
            anchor="w"
        )
        self.file_label.grid(row=1, column=1, padx=5, pady=5, sticky="ew")

        self.upload_button = tk.Button(
            self.root,
            text="Browse",
            command=self.upload_file
        )
        self.upload_button.grid(row=1, column=2, padx=5, pady=5, sticky="ew")

    def create_processing_section(self):
        """Create UI elements related to processing and providing the API key."""
        tk.Label(
            self.root,
            text="Please enter your OpenAI API key:",
            bg="white",
            fg="black",
            anchor="w"
        ).grid(row=2, column=0, padx=5, pady=5, sticky="ew")

        self.entry_api_key = tk.Entry(self.root, state=tk.DISABLED)
        self.entry_api_key.grid(row=2, column=1, padx=5, pady=5, sticky="ew")

        self.process_button = tk.Button(
            self.root,
            text="Process",
            command=self.start_processing,
            state=tk.DISABLED
        )
        self.process_button.grid(row=2, column=2, padx=5, pady=5, sticky="ew")

        self.process_log_label = tk.Label(
            self.root,
            text="Process Log",
            bg="white",
            fg="black",
            anchor="w"
        )
        self.process_log_label.grid(row=3, column=0, padx=5, pady=5, columnspan=2, sticky="ew")

    def create_abort_section(self):
        """Create UI elements to abort the ongoing process."""
        self.abort_button = tk.Button(
            self.root,
            text="Abort",
            command=self.request_abort,
            state=tk.DISABLED
        )
        self.abort_button.grid(row=4, column=2, padx=5, pady=5, sticky="ew")

    def download_template_file(self):
        """Save the default Excel template to the local directory."""
        save_excel_template("Template.xlsx")

    def upload_file(self):
        """Open a file dialog for the user to select an Excel file."""
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            self.input_file_path = file_path
            selected_file_name = os.path.basename(file_path)
            self.file_label.config(text=selected_file_name, fg="#4d5d53")
            self.process_button.config(state=tk.NORMAL)
            self.entry_api_key.config(state=tk.NORMAL)

    def start_processing(self):
        """Start the translation process if the API key is valid."""
        if not utils.is_valid_openai_api_key(self.entry_api_key.get()):
            messagebox.showerror("Error", "Invalid API key for OpenAI.")
            return

        self.api_key = self.entry_api_key.get()
        selected_file_name = self.file_label.cget("text")
        self.process_log_label.config(text=f"Processing has started for the file {selected_file_name}")

        # Disable inputs to prevent changes during processing
        self.entry_api_key.config(state=tk.DISABLED)
        self.upload_button.config(state=tk.DISABLED)
        self.abort_button.config(state=tk.NORMAL, highlightbackground="red")

        self.abort_requested_state = False

        # Run processing in a separate thread so the UI doesn't freeze
        self.processing_thread = threading.Thread(target=self.process_file_in_thread)
        self.processing_thread.start()

    def process_file_in_thread(self):
        """Run file processing in a separate thread to keep the UI responsive."""
        try:
            translated_workbook = self.process_file(
                file_path=self.input_file_path,
                api_key=self.api_key,
                stop_condition=lambda: self.abort_requested_state,
            )

            # If user requested abort, handle it gracefully
            if self.abort_requested_state:
                self.process_log_label.config(text="Processing was aborted by user.")
                return

            # If we got a translated workbook, prompt user to save it
            if translated_workbook:
                saved_path = utils.save_excel_workbook(translated_workbook)
                if saved_path:
                    self.process_log_label.config(text="Processing completed successfully.")
                else:
                    self.process_log_label.config(text="Save operation was canceled by the user.")
            else:
                # No data or aborted early without workbook
                if not self.abort_requested_state:
                    self.process_log_label.config(text="No data to save or processing ended unexpectedly.")

        except Exception as e:
            self.process_log_label.config(text=f"Error during processing: {e}")

        finally:
            # Restore UI state
            self.file_label.config(text="Please upload the excel file you want to translate:")
            self.upload_button.config(state=tk.NORMAL)
            self.process_button.config(state=tk.DISABLED)

            self.entry_api_key.config(state=tk.NORMAL)
            self.entry_api_key.delete(0, tk.END)
            self.entry_api_key.config(state=tk.DISABLED)

            self.abort_button.config(highlightbackground=None, state=tk.DISABLED)

    def process_file(self, file_path: str, api_key: str, stop_condition=None):
        """
        Process the Excel file, translating product descriptions from English to Turkish.
        Uses the provided API key for OpenAI requests. The `stop_condition` is a callable
        that returns True if processing should stop (e.g., user clicked 'Abort').
        """
        client = OpenAI(api_key=api_key)

        # Load the original workbook and read rows
        original_wb = openpyxl.load_workbook(file_path)
        original_sheet = original_wb.active

        rows_data = []
        for row in original_sheet.iter_rows(min_row=2, values_only=True):  # Skip headers
            product_code_col = row[0]
            product_description_col = row[1]
            if product_code_col is not None and product_description_col is not None:
                rows_data.append((product_code_col, product_description_col))

        # Prepare new workbook for translated data
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active
        new_sheet.title = 'TranslateToTurkishFromEnglish'
        new_sheet.cell(row=1, column=1).value = "Urun Kodu"
        new_sheet.cell(row=1, column=2).value = "Turkce Urun Aciklamasi"

        # Prompt template for translation
        prompt_template = """
        Translate the following English product description to Turkish. Format the translation in HTML as follows:

        <div>
          <h3><b>Ürün Açıklaması:</b></h3>
          <p>[Translated general description here. Provide a natural and easy-to-read translation.]</p>
        </div>

        <div>
          <h3><b>Ürün Özellikleri:</b></h3>
          <ul>
            <li>[Feature 1]</li>
            <li>[Feature 2]</li>
            <!-- Add more bullet points as needed -->
          </ul>
        </div>

        <div>
          <h3><b>Kompozisyon:</b></h3>
          <p>[Translated composition here. Keep to composition clean and concise.]</p>
        </div>

        Only include a section if the corresponding information exists in the input. Ensure the translation is contextually accurate and easy to understand.
        Only include the content inside HTML tags. Do not add any additional tags like "```html" at the beginning or "```" at the end.

        Input:
        {description}

        Output:
        """

        # Translate each row
        current_row = 2
        for product_code, english_description in rows_data:
            # Check if user requested abort
            if stop_condition and stop_condition():
                return None

            formatted_prompt = prompt_template.format(description=english_description)

            try:
                response = client.chat.completions.create(
                    model="gpt-4o",
                    messages=[
                        {"role": "system",
                         "content": "You are a helpful assistant that translates English to Turkish."},
                        {"role": "user", "content": formatted_prompt}
                    ],
                    temperature=0.2
                )
                translated_text = response.choices[0].message.content.strip()

            except Exception as e:
                print(f"Error translating text '{english_description[:30]}...': {e}")
                translated_text = "Error"

            new_sheet.cell(row=current_row, column=1).value = product_code
            new_sheet.cell(row=current_row, column=2).value = translated_text
            current_row += 1
            print(current_row)
        return new_wb

    def request_abort(self):
        self.abort_requested_state = True
        self.process_log_label.config(text="Abort requested... Attempting to stop.")
